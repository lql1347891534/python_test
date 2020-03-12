import openpyxl
from common.request import Request
import json
from common import logger

class Case:

    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None

class DoExcel:

    def __init__(self,file_name):
        self.file_name=file_name
        try:
            self.workbook=openpyxl.load_workbook(filename=file_name)
            logger.MyLog.info("open workbook in the path {0}".format(file_name))
        except FileNotFoundError as a:
            logger.MyLog.error("{0} not found, please chack file path".format(file_name))
            raise a

    def get_case(self,sheet_name):
        sheet=self.workbook[sheet_name]#获取到测试用例对应的sheet
        logger.MyLog.info("get sheet on workbook")
        max_row=sheet.max_row#获取最大行
        cases=[]
        for r in range(2,max_row+1):
            case=Case()#实例化一个case对象
            case.case_id=sheet.cell(row=r,column=1).value
            logger.MyLog.info("get case_id:{0} from excel".format(case.case_id))
            case.title=sheet.cell(row=r,column=2).value
            logger.MyLog.info("get title:{0} from excel".format(case.title))
            case.url=sheet.cell(row=r,column=3).value
            logger.MyLog.info("get url:{0} from excel".format(case.url))
            case.data=sheet.cell(row=r,column=4).value
            logger.MyLog.info("get data:{0} from excel".format(case.data))
            case.method=sheet.cell(row=r,column=5).value
            logger.MyLog.info("get method:{0} from excel".format(case.method))
            case.expected=sheet.cell(row=r,column=6).value
            logger.MyLog.info("get expected:{0} from excel".format(case.expected))
            cases.append(case)
        return cases

    def get_sheet_names(self):
        logger.MyLog.info("get all sheetnames")
        return self.workbook.sheetnames

    #根据sheetname定位到对应的sheet,根据case_id定位到当前行的actual,result
    # 将actual和result的值写入Excel中，并保存整个workbook
    def write_actual_by_case_id(self,sheetname,case_id,actual,result):
        sheet=self.workbook[sheetname]
        max_row=sheet.max_row
        for r in range(2,max_row+1):
            if case_id==sheet.cell(r,1).value:
                sheet.cell(r,7).value=actual
                logger.MyLog.info("write actual(0) in excel".format(actual))
                sheet.cell(r,8).value=result
                logger.MyLog.info("write result:{0} in excel".format(result))
                self.workbook.save(self.file_name)
                break

if __name__ == '__main__':
    # 实例化一个do_excel对象
    do_excel=DoExcel(r"D:\Python_test\datas\test_data.xlsx")
    # 获取workbook中所有的sheet
    # sheetnames=do_excel.get_sheet_names()
    # print("所有sheet的名称：{0}".format(sheetnames))
    # # 遍历所有的sheet，依次执行所有sheet对应的测试用例
    # for sheetname in sheetnames:
    #     print("当前执行sheet的名称：{0}".format(sheetname))
    #     # 调用方法获取sheet中的测试用例
    #     test_case=do_excel.get_case("login")
    #     # 循环取出Excel中对应的每条测试用例，并请求
    #     for case in test_case:
    #         # 使用json模块将Excel中的str类型反序列化成dict
    #         data=json.loads(case.data)
    #         # 调用封装的http请求方法，执行Excel中的测试用例
    #         resq=Request(method=case.method,url=case.url,data=data)
    #         # 将Excel中期望值是null的转化成python可识别的None类型
    #         if case.expected=="null":
    #             expected = json.loads(case.expected, encoding="utf-8")
    #         else:
    #             expected=case.expected
    #         print(resq.get_json()["data"]["token"]["access_token"])
    #         # 断言实际值与期望值是否一致
    #         # 将实际值写入Excel中
    #         # if resq.get_json()["msg"]==None:
    #         #     actual=json.dumps(resq.get_json()["msg"])
    #         # else:
    #         #     actual=resq.get_json()["msg"]
    #         # if resq.get_json()["msg"] == expected:
    #         #     do_excel.write_actual_by_case_id(sheetname=sheetname,case_id=case.case_id,
    #         #                                      actual=actual,result="pass")
    #         #     print("pass")
    #         # else:
    #         #     do_excel.write_actual_by_case_id(sheetname=sheetname, case_id=case.case_id,
    #         #                                      actual=actual, result="fail")
    #         #     print("fail")
    test_case=do_excel.get_case("reach")
    for case in test_case:
        print(case.data)